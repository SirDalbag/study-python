import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import json
import datetime


class Worker:
    BONUS = 1.15

    def __init__(
        self,
        name: str,
        position: str,
        salary: float,
        is_reserved: bool,
        date: datetime.datetime,
    ):
        self.name: str = name
        self.position: str = position
        self.salary: float = float(salary)
        self.is_reserved: bool = is_reserved
        self.date: datetime.datetime = date

    def __repr__(self) -> str:
        return f"<Worker> {self.name} - {self.position}"

    def serialize(self) -> dict:
        return {
            "name": self.name,
            "position": self.position,
            "salary": self.salary,
            "is_reserved": self.is_reserved,
            "date": self.date,
        }


def read():
    workbook: Workbook = openpyxl.load_workbook("less22/data.xlsx")
    worksheet: Worksheet = workbook.active

    workers = []
    for i in range(2, worksheet.max_row + 1):
        worker = Worker(
            name=worksheet.cell(row=i, column=1).value,
            position=worksheet.cell(row=i, column=2).value,
            salary=worksheet.cell(row=i, column=3).value,
            is_reserved=True
            if worksheet.cell(row=i, column=4).value == "да"
            else False,
            date=worksheet.cell(row=i, column=5).value,
        )
        workers.append(worker)
    return workers


if __name__ == "__main__":
    workers = read()

    for worker in workers:
        print(worker)
        print(worker.serialize())

    reserved = list(filter(lambda x: x.is_reserved, workers))

    print(reserved)
