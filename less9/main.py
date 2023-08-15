import re


# Task 1. Получение доменов из списка e-mail адресов.
def extract_domains(email_list: list[str]) -> list[str]:
    return [
        re.search(r"@([\w.-]+)", x).group(1)
        for x in email_list
        if re.search(r"@([\w.-]+)", x)
    ]


# Task 2. Получение слов, начинающихся на гласную букву.
def get_words(text: str) -> list[str]:
    return re.findall(r"\b[aeiouаеёиоуыэюя]\w*", text, re.IGNORECASE)


# Task 3. Разбиение строки.
def split_string(text: str, delimiters: str) -> list[str]:
    return re.split("|".join(map(re.escape, delimiters)), text)


# Task 4. Выборы.
candidates = ["Аскаров", "Бекмуханов", "Ернур", "Пешая", "Карим", "Шаримазданов"]
votes = []

while True:
    vote = input("Вы отдаете выбор за: ")
    if vote == "Стоп":
        break
    votes.append(vote)

votes = tuple(votes)

winner = max(set(votes), key=votes.count)

votes_count = votes.count(winner)

print("Победитель выборов:", winner)
print("Количество голосов победителя:", votes_count)

# Task 5. Объединение Exсel файлов.
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

workbook: Workbook = openpyxl.load_workbook("less9\ОбщийЛист.xlsx")
worksheet: Worksheet = workbook.active

files = ["less9/Лист1.xlsx", "less9/Лист2.xlsx", "less9/Лист3.xlsx"]

current_collumn = 1

for i in files:
    wb: Workbook = openpyxl.load_workbook(i)
    ws: Worksheet = wb.active

    for row in range(1, ws.max_row + 1):
        for column in range(current_collumn, current_collumn + 1):
            cell_value = ws.cell(row=row, column=column).value
            worksheet.cell(row=row, column=column).value = cell_value

    current_collumn += 1
    wb.close()

workbook.save("less9\ОбщийЛист.xlsx")
workbook.close()
