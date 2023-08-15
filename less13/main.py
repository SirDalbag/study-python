import sqlite3
import sys
from PyQt6.QtCore import QRect
from PyQt6.QtWidgets import (
    QApplication,
    QWidget,
    QGridLayout,
    QLabel,
    QLineEdit,
    QPushButton,
)
from openpyxl import load_workbook


def query(query_str: str, args=(), many=True) -> list | None:
    with sqlite3.connect("less13/database.db") as connection:
        cursor = connection.cursor()
        cursor.execute(query_str, args)
        try:
            if many:
                return cursor.fetchall()
            return cursor.fetchone()
        except Exception as error:
            return None


def select_db() -> list[tuple[any]] | None:
    return query("SELECT id, name, count, price FROM products", many=True)


def insert_db(name: str, count: int, price: float) -> None:
    query(
        "INSERT INTO products (name, count, price) VALUES (?, ?, ?);",
        (name, count, price),
    )


def get_data() -> list:
    wb = load_workbook(filename=f"less13/{str(file_name.text())}")
    ws = wb.active

    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    wb.close()
    return data


def import_data():
    for i in get_data():
        insert_db(i[0], i[1], i[2])


def export_data():
    wb = load_workbook(filename=f"less13/{str(file_name.text())}")
    ws = wb.active

    for row_index, row_data in enumerate(select_db(), start=1):
        ws.cell(row=row_index, column=1, value=row_data[0])
        ws.cell(row=row_index, column=2, value=row_data[1])
        ws.cell(row=row_index, column=3, value=row_data[2])
        ws.cell(row=row_index, column=4, value=row_data[3])

    wb.save(filename=f"less13/{str(file_name.text())}")


if __name__ == "__main__":
    app = QApplication(sys.argv)

    window = QWidget()
    window.setWindowTitle("Import - Export")
    window.setGeometry(QRect(200, 200, 400, 100))

    grid = QGridLayout()
    grid.setSpacing(4)

    label_title = QLabel("Filеname")
    grid.addWidget(label_title, 0, 1)

    file_name = QLineEdit()
    grid.addWidget(file_name, 0, 2)

    import_button = QPushButton("Import")
    import_button.clicked.connect(import_data)
    grid.addWidget(import_button, 0, 3)

    export_button = QPushButton("Export")
    export_button.clicked.connect(export_data)
    grid.addWidget(export_button, 0, 4)

    window.setLayout(grid)

    window.show()
    app.exec()
